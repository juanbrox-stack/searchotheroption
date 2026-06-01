"""
Gestor de Sustitutos Cecotec
- Sinstocks col K (índice 10) = SKU/Referencia Cecotec
- Tarifa Nacional T_AMZ: col REFERENCIA → cruce, col PVP PUB. = precio
- Tarifa Inter ES-FR/ES-IT etc: col REFERENCIA → cruce, col PVP PUB = precio
- Stock España: col Stock Operativo (col G)
- Stock FR/IT/DE/IT: col StockDisponible (col F)
- Sustituto: misma SUBFAMILIA, no desposicionado, PVP PUB ≤ original + 10€, stock > 0
- Salida: plantilla NUMBER / ARTICLE / NEW_ARTICLE
"""
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re, io, glob, tempfile
from pathlib import Path
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

st.set_page_config(page_title="Sustitutos Cecotec", page_icon="🔄", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Nunito:wght@600;700;800;900&family=Nunito+Sans:wght@400;600&display=swap');
:root{--blue:#3EB1C8;--black:#141413;--bg:#FAF9F5;--white:#fff;--grey:#e8e8e4;--muted:#6b7280;}
html,body,[class*="css"]{font-family:'Nunito Sans',sans-serif;background:var(--bg)!important;}
.stApp{background:var(--bg)!important;}
.navbar{background:var(--black);padding:.7rem 2rem;display:flex;align-items:center;gap:1.2rem;margin:-1rem -1rem 1.8rem -1rem;border-bottom:3px solid var(--blue);}
.navbar .logo{font-family:'Nunito',sans-serif;font-size:1.55rem;font-weight:900;color:#fff;letter-spacing:-.5px;}
.navbar .logo span{color:var(--blue);}
.navbar .sub{font-size:.72rem;color:rgba(255,255,255,.55);font-weight:600;text-transform:uppercase;letter-spacing:.12em;border-left:1px solid rgba(255,255,255,.2);padding-left:1.1rem;}
.sec{font-family:'Nunito',sans-serif;font-size:1rem;font-weight:800;color:var(--black);text-transform:uppercase;letter-spacing:.08em;padding:.4rem 0 .4rem .7rem;border-left:4px solid var(--blue);margin:1.4rem 0 .9rem;background:linear-gradient(90deg,rgba(62,177,200,.07),transparent);}
.kpi-row{display:flex;gap:.8rem;margin-bottom:1.4rem;flex-wrap:wrap;}
.kpi{background:var(--white);border-radius:10px;padding:.9rem 1.2rem;box-shadow:0 1px 3px rgba(20,20,19,.08),0 0 0 1px var(--grey);flex:1;min-width:110px;position:relative;overflow:hidden;}
.kpi::after{content:'';position:absolute;bottom:0;left:0;right:0;height:3px;background:var(--blue);}
.kpi .val{font-size:1.9rem;font-weight:900;color:var(--blue);font-family:'Nunito',sans-serif;line-height:1;}
.kpi .lbl{font-size:.68rem;color:var(--muted);text-transform:uppercase;letter-spacing:.07em;margin-top:.3rem;font-weight:600;}
div[data-testid="stTabs"] button[aria-selected="true"]{color:var(--blue)!important;border-bottom-color:var(--blue)!important;}
div[data-testid="stButton"] button[kind="primary"]{background:var(--blue)!important;border-color:var(--blue)!important;color:#fff!important;font-weight:800;border-radius:6px;}
div[data-testid="stDataFrame"] thead th{background:var(--black)!important;color:#fff!important;font-weight:700;font-size:.78rem;text-transform:uppercase;}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="navbar"><div class="logo">ceco<span>tec</span></div><div class="sub">Gestor de sustitutos · Sin stock</div></div>', unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
VALID_COUNTRIES = {'ES','FR','IT','DE','PT','PL','NL','BE'}
AMAZON_KEYWORDS = ['amazon','turaco-amazon','mano-a-mano']
# Tarifa inter: country → (sheet_name, pvp_col, neto_col)
INTER_SHEET = {
    'FR': ('ES-FR', 'PVP PUB', 'NETO ES-FR'),
    'IT': ('ES-IT', 'PVP PUB', 'NETO ES-IT'),
    'DE': ('ES-DE', 'PVP PUB', 'NETO ES-DE'),
    'PT': ('PT',    'PVP PUB', 'NETO PT'),
    'PL': ('PL',    'PVP PUB', 'NETO PL'),
    'NL': ('NL',    'PVP PUB', 'NETO NL'),
    'BE': ('BE',    'PVP PUB', 'NETO BE'),
}
UPLOAD_DIR = Path(__file__).parent

# ── Normalise reference ────────────────────────────────────────────────────────
def norm_ref(v) -> str:
    """Normalize reference: strip leading zeros for numeric, preserve A-prefix refs."""
    s = str(v).strip()
    if re.match(r'^A\d{2}_\w+_\d+$', s):
        return s  # A01_EU01_106744 — keep as-is
    if re.match(r'^\d+\.?\d*$', s):
        return str(int(float(s)))  # strip leading zeros
    return s

# ── Data loaders ───────────────────────────────────────────────────────────────
@st.cache_data
def load_tarifa_nac(path):
    df = pd.read_excel(path, sheet_name='T_AMZ')
    df['REF'] = df['REFERENCIA'].apply(norm_ref)
    return df

@st.cache_data
def load_tarifa_inter(path):
    xl = pd.ExcelFile(path)
    out = {}
    for sh in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sh)
        if 'REFERENCIA' not in df.columns:
            continue
        df['REF'] = df['REFERENCIA'].apply(norm_ref)
        out[sh] = df
    return out

@st.cache_data
def load_stock(path):
    """Returns dict country → DataFrame.
    España: col G = Stock Operativo
    FR/IT/DE/IT: col F = StockDisponible
    """
    xl = pd.ExcelFile(path)
    SHEET_MAP = {'España':'ES','Alemania':'DE','Francia':'FR','Italia':'IT'}
    out = {}
    for sh in xl.sheet_names:
        country = SHEET_MAP.get(sh, sh)
        df = pd.read_excel(path, sheet_name=sh)
        df['REF'] = df['Referencia'].apply(norm_ref)
        # Normalise stock column name
        if 'Stock Operativo' in df.columns:       # España
            df['STOCK'] = pd.to_numeric(df['Stock Operativo'], errors='coerce').fillna(0)
        elif 'StockDisponible' in df.columns:      # FR/IT/DE
            df['STOCK'] = pd.to_numeric(df['StockDisponible'], errors='coerce').fillna(0)
        else:
            df['STOCK'] = 0
        out[country] = df
    return out

@st.cache_data
def load_listing(path):
    """SKU → set of countries from Amazon orders file (sku + ship-country cols)."""
    p = Path(path)
    sku_countries = {}
    if p.suffix.lower() in ('.txt','.tsv','.csv'):
        df = pd.read_csv(path, sep='\t', encoding='utf-8-sig', dtype=str)
        cols = [c.strip() for c in df.columns]; df.columns = cols
        if 'sku' in cols and 'ship-country' in cols:
            for _, row in df[['sku','ship-country']].dropna().iterrows():
                country = str(row['ship-country']).strip().upper()
                if country not in VALID_COUNTRIES: continue
                m = re.search(r'(\d{3,6})$', str(row['sku']))
                if m:
                    sku_countries.setdefault(str(int(m.group(1))), set()).add(country)
        elif 'SKU del vendedor' in cols:
            for sku in df['SKU del vendedor'].dropna():
                s = str(sku).strip()
                m = re.match(r'^([A-Z]{2})0*(\d+)$', s)
                if m and m.group(1) in VALID_COUNTRIES:
                    sku_countries.setdefault(str(int(m.group(2))), set()).add(m.group(1))
                else:
                    m2 = re.search(r'(\d{3,6})$', s)
                    if m2: sku_countries.setdefault(str(int(m2.group(1))), set()).add('ES')
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb['BBDD'] if 'BBDD' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True)); wb.close()
        for r in rows[1:]:
            pais = str(r[1] or '').strip().upper()
            ref_raw = str(r[3] or '').strip()
            m = re.match(r'^([A-Z]{2})(\d+)$', ref_raw)
            if m:
                country = m.group(1) if m.group(1) in VALID_COUNTRIES else pais
                ref_num = str(int(m.group(2)))
            else:
                m2 = re.search(r'(\d{3,6})$', ref_raw)
                ref_num = str(int(m2.group(1))) if m2 else ref_raw
                country = pais if pais in VALID_COUNTRIES else 'ES'
            if ref_num: sku_countries.setdefault(ref_num, set()).add(country)
    return sku_countries


@st.cache_data
def load_feed(path):
    """Feed web Cecotec: mpn → url. Col C=link, col M=mpn."""
    df = pd.read_excel(path, usecols=['link','mpn'], dtype=str)
    df = df.dropna(subset=['mpn','link'])
    # Normalize mpn same as norm_ref: strip leading zeros for numeric
    df['REF'] = df['mpn'].apply(norm_ref)
    return df.set_index('REF')['link'].to_dict()

def get_smtp():
    try:
        return (st.secrets["correo"]["servidor_smtp"],
                int(st.secrets["correo"]["puerto"]),
                st.secrets["correo"]["usuario"],
                st.secrets["correo"]["password"])
    except: return None

@st.cache_data
def load_remitentes(b: bytes):
    df = pd.read_excel(io.BytesIO(b))
    df.columns = [c.strip() for c in df.columns]
    return df

def send_cancelados_email(to: str, asunto: str, df_cancel: pd.DataFrame, smtp_cfg):
    server, port, user, pwd = smtp_cfg
    msg = MIMEMultipart("mixed")
    msg["Subject"] = asunto
    msg["From"]    = user
    msg["To"]      = to

    rows_html = "".join(
        f'<tr style="background:{"#f8fafc" if i%2==0 else "#fff"}">'
        + "".join(
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">{v}</td>'
            for v in [row.get("Expedición",""), row.get("Canal",""), row.get("SKU",""),
                      row.get("Producto",""), row.get("País",""), row.get("Motivo","")]
        ) + "</tr>"
        for i, (_, row) in enumerate(df_cancel.iterrows())
    )
    html = f"""<html><body style="font-family:Arial;color:#141413;max-width:800px;margin:0 auto">
    <div style="background:#141413;padding:18px 24px;border-radius:10px 10px 0 0;border-bottom:3px solid #3EB1C8">
      <h2 style="color:#fff;margin:0;font-size:19px">❌ {asunto}</h2>
    </div>
    <div style="padding:18px 24px;border:1px solid #e8e8e4;border-top:none;border-radius:0 0 10px 10px">
      <table style="width:100%;border-collapse:collapse;font-size:13px">
        <thead><tr style="background:#141413;color:#fff">
          <th style="padding:8px 12px;text-align:left">Expedición</th>
          <th style="padding:8px 12px;text-align:left">Canal</th>
          <th style="padding:8px 12px;text-align:left">SKU</th>
          <th style="padding:8px 12px;text-align:left">Producto</th>
          <th style="padding:8px 12px;text-align:left">País</th>
          <th style="padding:8px 12px;text-align:left">Motivo</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
      <p style="font-size:11px;color:#94a3b8;margin-top:16px">Generado por Gestor de Sustitutos Cecotec</p>
    </div></body></html>"""

    msg.attach(MIMEText(html, "html", "utf-8"))

    # Excel attachment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Cancelados"
    thin = Side(style="thin", color="FFD0D0CC")
    brd  = Border(top=thin, left=thin, right=thin, bottom=thin)
    headers = ["Expedición","Canal","SKU","Producto","País","PVP (€)","Subfamilia","Motivo"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", start_color="FF141413")
        c.alignment = Alignment(horizontal="center"); c.border = brd
    for ri, (_, row) in enumerate(df_cancel.iterrows(), 2):
        fill = PatternFill("solid", start_color="FFF5F5F3" if ri%2==0 else "FFFFFFFF")
        for ci, key in enumerate(["Expedición","Canal","SKU","Producto","País","PVP (€)","Subfamilia","Motivo"], 1):
            c = ws.cell(row=ri, column=ci, value=row.get(key,""))
            c.font = Font(name="Arial", size=10); c.fill = fill; c.border = brd
    for i, w in enumerate([15,25,14,40,6,10,20,35], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); xlsx_bytes = buf.getvalue()

    part = MIMEBase("application", "octet-stream")
    part.set_payload(xlsx_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", 'attachment; filename="cancelados.xlsx"')
    msg.attach(part)

    if int(port) == 465:
        with smtplib.SMTP_SSL(server, int(port)) as s:
            s.login(user, pwd); s.sendmail(user, to, msg.as_string())
    else:
        with smtplib.SMTP(server, int(port)) as s:
            s.ehlo(); s.starttls(); s.ehlo()
            s.login(user, pwd); s.sendmail(user, to, msg.as_string())

def parse_sinstocks(path):
    """
    sinstock.xlsx — headers in row 0, data from row 1.
    Uses openpyxl to avoid pandas column shift issues with empty cells.
    Col J (idx 9)  = ARTÍCULO  → "05993 - NOMBRE" or "A01_EU01_106744 - NOMBRE"
    Col E (idx 4)  = EXPEDICIÓN (data is one col right of header label)
    Col N (idx 13) = ENTIDAD
    Col S (idx 18) = CÓDIGO DE PAÍS
    """
    import unicodedata
    def strip_acc(s):
        return ''.join(c for c in unicodedata.normalize('NFD', str(s))
                       if unicodedata.category(c) != 'Mn').upper().strip()

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return pd.DataFrame()

    # Build header map: normalized_name → col_index
    hdr = {strip_acc(str(v or '')): i for i, v in enumerate(rows[0]) if v}

    def gc(fragments, default):
        """Get column index by fragment match."""
        for frag in fragments:
            for k, idx in hdr.items():
                if frag in k:
                    return idx
        return default

    # Note: EXPEDICIÓN header is at col 3 but DATA is at col 4 (one right)
    # This is because col D header = EXPEDICIÓN but col D data = None, col E data = value
    # We detect this by checking which adjacent col actually has D26... values
    expedi_hdr_idx = gc(['EXPEDICION','EXPEDICI'], 3)
    # Check if data is in header col or one to the right
    sample_vals = [str(rows[i][expedi_hdr_idx] or '') for i in range(1, min(5, len(rows)))]
    if not any(v.startswith('D') for v in sample_vals):
        expedi_hdr_idx += 1  # shift right

    c_expedi   = expedi_hdr_idx
    c_articulo = gc(['ARTICULO'], 9)
    c_cantidad = gc(['CANTIDAD'], 8)
    c_entidad  = gc(['ENTIDAD'], 13)
    c_pais     = gc(['CODIGO DE PAIS', 'PAIS'], 18)
    c_order_id = gc(['IDENTIFICADOR'], 1)

    records = []
    for r in rows[1:]:
        expedi = str(r[c_expedi] or '').strip()
        if not expedi or expedi in ('None', 'nan', ''):
            continue

        articulo = str(r[c_articulo] or '').strip()
        if not articulo or articulo in ('None', 'nan', ''):
            continue

        # Parse SKU and name
        m_num = re.match(r'^(\d{3,6})\s*[-–]\s*(.*)$', articulo)
        m_axx = re.match(r'^(A\d{2}_\w+_\d+)\s*[-–]\s*(.*)$', articulo)

        if m_num:
            sku  = str(int(m_num.group(1)))
            name = m_num.group(2).strip()
        elif m_axx:
            sku  = m_axx.group(1).strip()
            name = m_axx.group(2).strip()
        else:
            sku  = articulo
            name = articulo

        marketplace = str(r[c_entidad] or '').strip()
        country     = str(r[c_pais] or '').strip().upper()
        if country not in VALID_COUNTRIES:
            country = 'ES'

        is_amazon = any(k in marketplace.lower() for k in AMAZON_KEYWORDS)

        try:    qty = int(float(str(r[c_cantidad] or 1)))
        except: qty = 1

        records.append({
            'EXPEDICION':  expedi,
            'SKU':         sku,
            'NOMBRE':      name[:80],
            'CANTIDAD':    qty,
            'PAIS':        country,
            'MARKETPLACE': marketplace,
            'IS_AMAZON':   is_amazon,
            'ORDER_ID':    r[c_order_id],
        })
    return pd.DataFrame(records)

# ── Tarifa lookup ──────────────────────────────────────────────────────────────
def get_tarifa(sku, country, is_amazon, tarifa_nac, tarifa_inter):
    """Returns (row, df, pvp_col, neto_col) for a given SKU and context."""
    if is_amazon and country != 'ES' and country in INTER_SHEET:
        sh, pvp_col, neto_col = INTER_SHEET[country]
        df = tarifa_inter.get(sh, tarifa_nac)
        # Fallback col names
        if pvp_col not in df.columns:
            pvp_col = next((c for c in df.columns if 'PVP PUB' in str(c).upper()), pvp_col)
        if neto_col not in df.columns:
            neto_col = next((c for c in df.columns if 'NETO' in str(c).upper()), 'NETO')
    else:
        df = tarifa_nac
        pvp_col  = 'PVP PUB.'
        neto_col = 'NETO'

    mask = df['REF'] == str(sku)
    if not mask.any():
        return None, df, pvp_col, neto_col
    return df[mask].iloc[0], df, pvp_col, neto_col

def get_stock_for_ref(sku, country, is_amazon, stocks):
    search = [country, 'ES'] if (is_amazon and country in stocks) else ['ES']
    for c in search:
        df = stocks.get(c)
        if df is None: continue
        mask = df['REF'] == str(sku)
        if mask.any():
            return int(df[mask].iloc[0]['STOCK'] or 0)
    return 0

def extract_size(text: str) -> str | None:
    """Extract dimensions like 140x190, 80x200, 135x190 from product name."""
    m = re.search(r'\b(\d{2,3})[xX×](\d{2,3})\b', str(text))
    if m:
        return f"{m.group(1)}x{m.group(2)}"
    return None

# Product families where size must match
SIZE_SENSITIVE_FAMILIES = {
    'colchon', 'colchones', 'base', 'bases', 'canape', 'canapes',
    'somier', 'somiers', 'canapé', 'canapés',
}

def find_substitutes(tar_row, pvp_col, pvp_orig, sku_orig,
                     country, is_amazon, tarifa_nac, tarifa_inter, stocks, max_extra=10.0):
    """Same SUBFAMILIA, not desposicionado, PVP PUB ≤ original+max_extra, stock > 0."""
    if is_amazon and country != 'ES' and country in INTER_SHEET:
        sh, pc, nc = INTER_SHEET[country]
        df = tarifa_inter.get(sh, tarifa_nac).copy()
        if pc not in df.columns:
            pc = next((c for c in df.columns if 'PVP PUB' in str(c).upper()), pc)
        stock_country = country
    else:
        df = tarifa_nac.copy()
        pc = 'PVP PUB.'
        stock_country = 'ES'

    subfamilia = str(tar_row.get('SUBFAMILIA',''))
    familia    = str(tar_row.get('FAMILIA',''))
    nombre_orig_full = str(tar_row.get('NOMBRE COMPLETO',''))

    # Extract size from original product name (e.g. 140x190)
    orig_size = extract_size(nombre_orig_full)
    # Check if this family requires size matching
    needs_size_match = orig_size and any(
        s in subfamilia.lower() or s in familia.lower()
        for s in SIZE_SENSITIVE_FAMILIES
    )

    # Filter by subfamilia (or familia fallback)
    if subfamilia and 'SUBFAMILIA' in df.columns:
        mask = df['SUBFAMILIA'].str.lower() == subfamilia.lower()
    elif familia and 'FAMILIA' in df.columns:
        mask = df['FAMILIA'].str.lower() == familia.lower()
    else:
        return pd.DataFrame()

    # Not desposicionado
    if 'DESPOSICIONADO' in df.columns:
        mask &= df['DESPOSICIONADO'].fillna(False) == False

    # PVP >= original (no sustituto de menos valor) AND <= original + max_extra
    if pc in df.columns:
        pvp_vals = pd.to_numeric(df[pc], errors='coerce')
        mask &= pvp_vals >= pvp_orig
        mask &= pvp_vals <= (pvp_orig + max_extra)

    # Size filter for mattresses and similar (must match dimensions exactly)
    if needs_size_match and 'NOMBRE COMPLETO' in df.columns:
        mask &= df['NOMBRE COMPLETO'].apply(
            lambda n: extract_size(str(n)) == orig_size
        )

    # Exclude original SKU
    mask &= df['REF'] != str(sku_orig)

    candidates = df[mask].copy()
    if candidates.empty:
        return pd.DataFrame()

    # Check stock
    stock_df  = stocks.get(stock_country, stocks.get('ES', pd.DataFrame()))
    stock_map = {} if stock_df.empty else stock_df.set_index('REF')['STOCK'].to_dict()

    results = []
    for _, row in candidates.iterrows():
        ref_c = str(row.get('REF',''))
        stk   = int(float(stock_map.get(ref_c, 0) or 0))
        if stk > 0:
            pvp_c  = float(pd.to_numeric(row.get(pc, 0), errors='coerce') or 0)
            neto_c = float(pd.to_numeric(row.get('NETO' if 'NETO' in row.index else pc, 0), errors='coerce') or 0)
            results.append({
                'REFERENCIA':      row.get('REFERENCIA',''),
                'REF':             ref_c,
                'NOMBRE COMPLETO': row.get('NOMBRE COMPLETO',''),
                'SUBFAMILIA':      row.get('SUBFAMILIA',''),
                'PVP':             pvp_c,
                'ΔPVP':            round(pvp_c - pvp_orig, 2),
                'STOCK':           stk,
            })

    if not results:
        return pd.DataFrame()
    return pd.DataFrame(results).sort_values(['ΔPVP','PVP'])

# ── Excel output ───────────────────────────────────────────────────────────────
def build_regen_excel(output_rows):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Regeneracion'
    thin = Side(style='thin', color='FFD0D0CC')
    brd  = Border(top=thin, left=thin, right=thin, bottom=thin)
    for ci, h in enumerate(['NUMBER','ARTICLE','NEW_ARTICLE'], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
        c.fill = PatternFill('solid', start_color='FF141413')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = brd
    ws.row_dimensions[1].height = 22
    for ri, row in enumerate(output_rows, 2):
        fill = PatternFill('solid', start_color='FFF5F5F3' if ri%2==0 else 'FFFFFFFF')
        for ci, val in enumerate([row['NUMBER'], row['ARTICLE'], row['NEW_ARTICLE']], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name='Arial', size=10); c.fill = fill
            c.border = brd; c.alignment = Alignment(vertical='center')
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── File helpers ───────────────────────────────────────────────────────────────
def _find(key, *patterns):
    if f'path_{key}' in st.session_state:
        p = st.session_state[f'path_{key}']
        if Path(p).exists(): return p
    for pat in patterns:
        m = glob.glob(str(UPLOAD_DIR / f'*{pat}*'))
        if m: return m[0]
    return None

# ── SIDEBAR uploaders ──────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📂 Cargar ficheros")
    fu_nac     = st.file_uploader('Tarifa Nacional',       type=['xlsx'], key='u_nac')
    fu_inter   = st.file_uploader('Tarifa Internacional',  type=['xlsx'], key='u_inter')
    fu_stock   = st.file_uploader('Stock Global',          type=['xlsx'], key='u_stock')
    fu_sins    = st.file_uploader('Sinstocks',             type=['xlsx'], key='u_sins')
    fu_listing = st.file_uploader('Listing Amazon',        type=['xlsx','txt','tsv','csv'], key='u_list')
    fu_feed    = st.file_uploader('Feed Web Cecotec',        type=['xlsx'], key='u_feed')
    fu_remit   = st.file_uploader('Remitentes (Canal/Email)', type=['xlsx'], key='u_remit')

    if st.button('💾 Guardar', type='primary', use_container_width=True):
        tmp = Path(tempfile.mkdtemp()); saved = 0
        for key, fu in [('nac',fu_nac),('inter',fu_inter),('stock',fu_stock),
                        ('sins',fu_sins),('listing',fu_listing),('feed',fu_feed)]:
            if fu:
                p = tmp / fu.name; p.write_bytes(fu.read())
                st.session_state[f'path_{key}'] = str(p); saved += 1
        if fu_remit:
            st.session_state['remitentes_bytes'] = fu_remit.read()
            saved += 1
        for fn in [load_tarifa_nac, load_tarifa_inter, load_stock, load_listing, load_feed]:
            fn.clear()
        st.success(f'✅ {saved} fichero(s) guardados')
        st.rerun()
    st.divider()
    st.caption('O coloca los ficheros en la carpeta del proyecto.')

# ── Resolve paths ──────────────────────────────────────────────────────────────
path_nac     = _find('nac',     'TARIFA_NACIONAL')
path_inter   = _find('inter',   'TARIFA_TURACO')
path_stock   = _find('stock',   'Stock_Global')
path_sins    = _find('sins',    'Sinstocks', 'sinstocks')
path_listing = _find('listing', 'Informe_', 'listing', 'Listing')
path_feed    = _find('feed',    'feed_Espan', 'feed_espan', 'feed_')
if not path_listing:
    hits = glob.glob(str(UPLOAD_DIR/'*.txt')) + glob.glob(str(UPLOAD_DIR/'*listing*'))
    if hits: path_listing = hits[0]
if not path_feed:
    hits = glob.glob(str(UPLOAD_DIR/'*feed*'))
    if hits: path_feed = hits[0]

# Status
cs = st.columns(6)
for col, lbl, pth in zip(cs,
    ['Tarifa Nac.','Tarifa Inter','Stock Global','Sinstocks','Listing Amazon','Feed Web'],
    [path_nac, path_inter, path_stock, path_sins, path_listing, path_feed]):
    with col:
        (st.success if pth else st.warning)(f'{"✅" if pth else "⚠️"} {lbl}')

data_ok = all([path_nac, path_inter, path_stock, path_sins])

# ── TABS ───────────────────────────────────────────────────────────────────────
tab_proc, tab_manual, tab_res = st.tabs(['⚙️ Procesar', '🔍 Buscar SKU', '📊 Resultados'])

with tab_proc:
    if not data_ok:
        st.warning('👈 Sube los ficheros en el panel izquierdo para continuar.')
    else:
        tarifa_nac   = load_tarifa_nac(path_nac)
        tarifa_inter = load_tarifa_inter(path_inter)
        stocks       = load_stock(path_stock)
        listing_map  = load_listing(path_listing) if path_listing else {}
        feed_map     = load_feed(path_feed) if path_feed else {}

        st.markdown('<div class="sec">⚙️ Configuración</div>', unsafe_allow_html=True)
        ca, cb, cc = st.columns(3)
        with ca: max_extra       = st.number_input('Máx. incremento PVP sobre original (€)', 0.0, 50.0, 10.0, 1.0)
        with cb: default_country = st.selectbox('País por defecto', ['ES','FR','IT','DE','PT','PL'], index=0)
        with cc: skip_dupes      = st.checkbox('Ignorar DUPLICADO', value=True)

        if st.button('🚀 Procesar pedidos sin stock', type='primary', use_container_width=True):
            with st.spinner('Leyendo Sinstocks…'):
                df_sins = parse_sinstocks(path_sins)
                if df_sins.empty or 'EXPEDICION' not in df_sins.columns:
                    st.error('No se pudieron leer pedidos del fichero Sinstocks.')
                    try:
                        df_dbg = pd.read_excel(path_sins, nrows=2)
                        st.caption(f"Columnas detectadas: {list(df_dbg.columns)}")
                        st.caption(f"Primera fila: {df_dbg.iloc[0].to_dict() if len(df_dbg)>0 else 'vacío'}")
                    except Exception as e:
                        st.caption(f"Error: {e}")
                    st.stop()
                if skip_dupes:
                    df_sins = df_sins[~df_sins['EXPEDICION'].str.upper().str.startswith('DUPLICADO')]
                df_sins = df_sins[df_sins['EXPEDICION'].str.startswith('D')].copy()
                # Enrich country from listing for Amazon orders
                if listing_map:
                    def enrich(row):
                        if row['IS_AMAZON'] and row['PAIS'] == 'ES':
                            ctrs = listing_map.get(str(row['SKU']), set()) - {'ES'}
                            if len(ctrs) == 1: return list(ctrs)[0]
                        return row['PAIS']
                    df_sins['PAIS'] = df_sins.apply(enrich, axis=1)
                df_sins['PAIS'] = df_sins['PAIS'].fillna(default_country)

            prog = st.progress(0)
            total = len(df_sins)
            results = []

            for i, (_, row) in enumerate(df_sins.iterrows()):
                prog.progress((i+1)/total,
                    text=f"[{i+1}/{total}] {row['EXPEDICION']} · {row['NOMBRE'][:40]}")

                sku     = row['SKU']
                country = str(row['PAIS'])
                is_amz  = bool(row['IS_AMAZON'])

                tar_row, df_tar, pvp_col, neto_col = get_tarifa(
                    sku, country, is_amz, tarifa_nac, tarifa_inter)

                pvp_orig   = 0.0
                subfamilia = ''
                nombre_orig= row['NOMBRE']

                if tar_row is not None:
                    try: pvp_orig = float(pd.to_numeric(tar_row.get(pvp_col,0), errors='coerce') or 0)
                    except: pass
                    subfamilia  = str(tar_row.get('SUBFAMILIA',''))
                    nombre_orig = str(tar_row.get('NOMBRE COMPLETO', row['NOMBRE']))

                subs = pd.DataFrame()
                if tar_row is not None and pvp_orig > 0:
                    subs = find_substitutes(tar_row, pvp_col, pvp_orig, sku,
                                            country, is_amz, tarifa_nac, tarifa_inter,
                                            stocks, max_extra)

                results.append({
                    'row':         row.to_dict(),
                    'tar_row':     tar_row,
                    'pvp_col':     pvp_col,
                    'pvp_orig':    pvp_orig,
                    'subfamilia':  subfamilia,
                    'nombre_orig': nombre_orig,
                    'sku':         sku,
                    'subs':        subs,
                })

            prog.empty()
            st.session_state['results'] = results
            st.session_state['selections'] = {}  # reset selections for new results
            st.toast('✅ Procesado completado', icon='🎯')
            st.components.v1.html("""
            <script>
            function click(){const t=window.parent.document.querySelectorAll('[data-baseweb="tab"]');
            for(const x of t){if(x.textContent.includes('Resultados')){x.click();return true;}}return false;}
            let n=0,iv=setInterval(()=>{if(click()||++n>20)clearInterval(iv);},150);
            </script>""", height=0)
            st.rerun()

with tab_manual:
    if not data_ok:
        st.warning('👈 Carga los ficheros en el panel izquierdo para continuar.')
    else:
        st.markdown('<div class="sec">🔍 Buscar sustituto por SKU</div>', unsafe_allow_html=True)
        st.caption("Introduce un SKU/referencia Cecotec para encontrar sus posibles sustitutos con stock.")

        col_a, col_b, col_c = st.columns([2, 1, 1])
        with col_a:
            manual_sku = st.text_input("SKU / Referencia", placeholder="ej: 8425  o  A01_EU01_106744")
        with col_b:
            manual_country = st.selectbox("País / Tarifa", ['ES','FR','IT','DE','PT','PL'], key="m_country")
            manual_amazon  = st.checkbox("Pedido Amazon", value=False, key="m_amazon")
        with col_c:
            manual_max = st.number_input("Máx. incremento PVP (€)", 0.0, 50.0, 10.0, 1.0, key="m_max")

        if st.button("🔍 Buscar sustitutos", type="primary", use_container_width=True, key="btn_manual_search"):
            if not manual_sku.strip():
                st.error("Introduce un SKU.")
            else:
                sku = norm_ref(manual_sku.strip())
                # Load data (already cached)
                tarifa_nac   = load_tarifa_nac(path_nac)
                tarifa_inter = load_tarifa_inter(path_inter)
                stocks       = load_stock(path_stock)
                feed_map     = load_feed(path_feed) if path_feed else {}

                tar_row, df_tar, pvp_col, neto_col = get_tarifa(
                    sku, manual_country, manual_amazon, tarifa_nac, tarifa_inter)

                if tar_row is None:
                    st.error(f"SKU `{sku}` no encontrado en la tarifa {'internacional ' + manual_country if manual_amazon and manual_country != 'ES' else 'nacional'}.")
                else:
                    try:
                        pvp_orig = float(pd.to_numeric(tar_row.get(pvp_col, 0), errors='coerce') or 0)
                    except:
                        pvp_orig = 0.0
                    subfamilia  = str(tar_row.get('SUBFAMILIA',''))
                    nombre_orig = str(tar_row.get('NOMBRE COMPLETO', sku))
                    url_orig    = feed_map.get(str(sku), '')

                    # Product info
                    st.success(f"✅ **{nombre_orig}** · Subfamilia: {subfamilia} · PVP PUB: **{pvp_orig:.2f}€**")
                    if url_orig:
                        st.markdown(f"[🔗 Ver en cecotec.es]({url_orig})")

                    # Find substitutes
                    subs = find_substitutes(tar_row, pvp_col, pvp_orig, sku,
                                            manual_country, manual_amazon,
                                            tarifa_nac, tarifa_inter, stocks, manual_max)

                    if subs.empty:
                        st.warning(f"No hay sustitutos con stock en subfamilia **{subfamilia}** con PVP entre {pvp_orig:.2f}€ y {pvp_orig + manual_max:.2f}€.")
                    else:
                        st.markdown(f"**{len(subs)} sustituto(s) encontrado(s):**")

                        # Enrich with URL
                        subs_display = subs[['REFERENCIA','NOMBRE COMPLETO','PVP','ΔPVP','STOCK']].copy()
                        subs_display['URL'] = subs['REF'].apply(lambda r: feed_map.get(str(r), ''))
                        subs_display = subs_display.rename(columns={
                            'NOMBRE COMPLETO': 'Nombre',
                            'PVP':             'PVP PUB (€)',
                            'ΔPVP':            'Δ PVP (€)',
                        })

                        st.dataframe(subs_display, use_container_width=True, hide_index=True,
                            column_config={
                                'PVP PUB (€)': st.column_config.NumberColumn(format='%.2f €'),
                                'Δ PVP (€)':   st.column_config.NumberColumn(format='%.2f €'),
                                'URL':         st.column_config.LinkColumn('URL', display_text='🔗 Ver'),
                            })

                        # Download
                        st.download_button(
                            "⬇️ Descargar CSV",
                            subs_display.drop(columns=['URL']).to_csv(index=False).encode(),
                            f"sustitutos_{sku}.csv", "text/csv"
                        )

with tab_res:
    if 'results' not in st.session_state:
        st.info('Ejecuta el procesado en ⚙️ Procesar.')
        st.stop()

    # Guard: if results are from an old run with different schema, clear them
    results = st.session_state['results']
    if results and 'SKU' not in results[0]['row'] and 'REF' not in results[0]['row']:
        del st.session_state['results']
        st.warning('Resultados anteriores incompatibles. Vuelve a procesar.')
        st.stop()
    con_sust = [r for r in results if not r['subs'].empty]
    sin_sust = [r for r in results if r['subs'].empty]
    total    = len(results)

    st.markdown(f"""<div class="kpi-row">
      <div class="kpi"><div class="val">{total}</div><div class="lbl">Pedidos sin stock</div></div>
      <div class="kpi"><div class="val">{len(con_sust)}</div><div class="lbl">Con sustituto ✅</div></div>
      <div class="kpi"><div class="val">{len(sin_sust)}</div><div class="lbl">A cancelar ❌</div></div>
      <div class="kpi"><div class="val">{len(con_sust)*100//total if total else 0}%</div><div class="lbl">Tasa resolución</div></div>
    </div>""", unsafe_allow_html=True)

    t1, t2 = st.tabs(['✅ Con sustituto', '❌ A cancelar'])

    with t1:
        if not con_sust:
            st.info('No hay pedidos con sustituto disponible.')
        else:
            table_rows, output_rows = [], []
            for r in con_sust:
                row  = r['row']
                expedi   = row['EXPEDICION']
                sel_idx  = st.session_state.get('selections', {}).get(expedi, 0)
                sel_idx  = min(sel_idx, len(r['subs'])-1)
                best     = r['subs'].iloc[sel_idx]
                delta = float(best.get('ΔPVP', 0) or 0)
                ref_orig = row.get('SKU') or row.get('REF','')
                ref_sust = str(best.get('REF',''))
                url_orig = feed_map.get(str(ref_orig), '')
                url_sust = feed_map.get(ref_sust, '')
                table_rows.append({
                    'Expedición':        row['EXPEDICION'],
                    'País':              row['PAIS'],
                    'Canal':             row['MARKETPLACE'][:30],
                    'Amazon':            '✅' if row['IS_AMAZON'] else '—',
                    'SKU original':      ref_orig,
                    'Producto original': r['nombre_orig'][:55],
                    'URL original':      url_orig,
                    'PVP orig. (€)':     round(r['pvp_orig'], 2),
                    'Subfamilia':        r['subfamilia'],
                    '→ SKU sustituto':   str(best.get('REFERENCIA','')),
                    '→ Sustituto':       str(best.get('NOMBRE COMPLETO',''))[:55],
                    'Medida':            extract_size(str(best.get('NOMBRE COMPLETO',''))) or '—',
                    'PVP sust. (€)':     round(float(best.get('PVP',0) or 0), 2),
                    'Δ PVP (€)':         delta,
                    'Stock disponible':  int(best.get('STOCK',0) or 0),
                    'URL sustituto':     url_sust,
                })
                output_rows.append({
                    'NUMBER':      row['EXPEDICION'],
                    'ARTICLE':     ref_orig,
                    'NEW_ARTICLE': str(best.get('REFERENCIA','')),
                })

            df_t = pd.DataFrame(table_rows)

            # Note if there are custom selections
            selections = st.session_state.get('selections', {})
            if any(v > 0 for v in selections.values()):
                st.info(f"⚠️ Tienes **{sum(1 for v in selections.values() if v > 0)} sustitutos modificados** manualmente. La tabla y descarga ya reflejan tu selección.")

            st.dataframe(df_t, use_container_width=True, hide_index=True, column_config={
                'PVP orig. (€)': st.column_config.NumberColumn(format='%.2f €'),
                'PVP sust. (€)': st.column_config.NumberColumn(format='%.2f €'),
                'Δ PVP (€)':     st.column_config.NumberColumn(format='%.2f €'),
                'URL original':  st.column_config.LinkColumn('URL original',  display_text='🔗 Ver'),
                'URL sustituto': st.column_config.LinkColumn('URL sustituto', display_text='🔗 Ver'),
            })

            dc1, dc2 = st.columns(2)
            with dc1:
                st.download_button('⬇️ Plantilla regeneración Excel',
                    build_regen_excel(output_rows), 'regeneracion_pedidos.xlsx',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True, type='primary')
            with dc2:
                st.download_button('⬇️ CSV sustitutos',
                    df_t.to_csv(index=False).encode(), 'sustitutos.csv',
                    'text/csv', use_container_width=True)

            st.markdown('<div class="sec">🔎 Detalle y selección de sustituto por pedido</div>', unsafe_allow_html=True)
            scol1, scol2 = st.columns([5,1])
            with scol1:
                st.caption("Expande cada pedido para cambiar el sustituto sugerido. La tabla y descarga se actualizan automáticamente.")
            with scol2:
                if st.button("↺ Reset selecciones", key="reset_sel"):
                    st.session_state['selections'] = {}
                    st.rerun()

            # Track selections (default = first substitute)
            if 'selections' not in st.session_state:
                st.session_state['selections'] = {}

            for idx, r in enumerate(con_sust):
                row  = r['row']
                best = r['subs'].iloc[0]
                expedi   = row['EXPEDICION']
                ref_orig = row.get('SKU') or row.get('REF','')

                # Get current selection for this order (default = index 0)
                sel_idx  = st.session_state['selections'].get(expedi, 0)
                sel_idx  = min(sel_idx, len(r['subs'])-1)
                selected = r['subs'].iloc[sel_idx]

                delta = float(selected.get('ΔPVP',0) or 0)
                sign  = f'+{delta:.2f}€' if delta>0 else (f'{delta:.2f}€' if delta<0 else '±0')
                url_orig = feed_map.get(str(ref_orig), '')
                url_sust = feed_map.get(str(selected.get('REF','')), '')

                with st.expander(
                    f"**{expedi}** · {row['PAIS']} "
                    f"{'🛒' if row['IS_AMAZON'] else ''} · "
                    f"SKU {ref_orig} → **{selected.get('REFERENCIA','')}** ({sign} PVP)"
                ):
                    cc1, cc2 = st.columns(2)
                    with cc1:
                        st.markdown('**📦 Pedido original**')
                        st.markdown(f"Expedición: `{expedi}`")
                        st.markdown(f"Canal: `{row['MARKETPLACE']}`")
                        st.markdown(f"SKU: `{ref_orig}` · País: **{row['PAIS']}**")
                        st.markdown(f"**{r['nombre_orig']}**")
                        st.markdown(f"PVP PUB: **{r['pvp_orig']:.2f}€** · Subfamilia: {r['subfamilia']}")
                        if url_orig:
                            st.markdown(f"[🔗 Ver producto en cecotec.es]({url_orig})")
                    with cc2:
                        st.markdown('**🔄 Sustitutos con stock**')

                        # Build options for radio selector
                        subs_df = r['subs'].reset_index(drop=True)
                        options = []
                        for _, sub in subs_df.iterrows():
                            pvp_s = float(sub.get('PVP',0) or 0)
                            dp    = float(sub.get('ΔPVP',0) or 0)
                            sign_s = f'+{dp:.2f}€' if dp>0 else (f'{dp:.2f}€' if dp<0 else '±0')
                            options.append(
                                f"{sub.get('REFERENCIA','')} · {str(sub.get('NOMBRE COMPLETO',''))[:40]} · {pvp_s:.2f}€ ({sign_s}) · Stock: {int(sub.get('STOCK',0))}"
                            )

                        chosen = st.radio(
                            "Selecciona sustituto:",
                            options,
                            index=sel_idx,
                            key=f"radio_{expedi}_{idx}",
                            label_visibility="collapsed",
                        )
                        new_idx = options.index(chosen)
                        if new_idx != sel_idx:
                            st.session_state['selections'][expedi] = new_idx
                            st.rerun()

                        # Show selected details
                        selected_row = subs_df.iloc[new_idx]
                        url_sel = feed_map.get(str(selected_row.get('REF','')), '')
                        st.markdown(f"✅ **Seleccionado: {selected_row.get('REFERENCIA','')}** — {str(selected_row.get('NOMBRE COMPLETO',''))[:50]}")
                        if url_sel:
                            st.markdown(f"[🔗 Ver en cecotec.es]({url_sel})")

    with t2:
        if not sin_sust:
            st.success('✅ Todos los pedidos tienen sustituto disponible.')
        else:
            cancel_rows = []
            for r in sin_sust:
                row = r['row']
                sku = row.get('SKU') or row.get('REF') or ''
                motivo = ('SKU no encontrado en tarifa' if r['tar_row'] is None
                          else f"Sin sustituto con stock en subfamilia '{r['subfamilia']}'")
                cancel_rows.append({
                    'Expedición':   row.get('EXPEDICION',''),
                    'País':         row.get('PAIS',''),
                    'Canal':        str(row.get('MARKETPLACE',''))[:35],
                    'Amazon':       '✅' if row.get('IS_AMAZON') else '—',
                    'SKU':          sku,
                    'Producto':     r['nombre_orig'][:60],
                    'PVP (€)':      round(r['pvp_orig'],2),
                    'Subfamilia':   r['subfamilia'],
                    'Motivo':       motivo,
                })
            df_c = pd.DataFrame(cancel_rows)
            st.warning(f"⚠️ **{len(df_c)} pedidos** sin sustituto — deben cancelarse.")
            st.dataframe(df_c, use_container_width=True, hide_index=True,
                column_config={'PVP (€)': st.column_config.NumberColumn(format='%.2f €')})

            # ── Descarga CSV ──────────────────────────────────────────────────
            st.download_button('⬇️ CSV cancelaciones',
                df_c.to_csv(index=False).encode(), 'cancelaciones.csv',
                'text/csv', use_container_width=True)

            # ── Email widget ──────────────────────────────────────────────────
            st.markdown("---")
            st.markdown('<div class="sec">📧 Enviar cancelaciones por email</div>', unsafe_allow_html=True)

            smtp_cfg      = get_smtp()
            remit_bytes   = st.session_state.get('remitentes_bytes')
            remitentes_df = load_remitentes(remit_bytes) if remit_bytes else None

            if not smtp_cfg:
                st.info("⚠️ SMTP no configurado. Añade `[correo]` en **Streamlit Secrets** para enviar emails:\n```toml\n[correo]\nservidor_smtp = \"smtp.gmail.com\"\npuerto = 587\nusuario = \"tu@email.com\"\npassword = \"tu_password\"\n```")
            if remitentes_df is None:
                st.info("Sube el fichero de **Remitentes** (columnas: Canal, Email, Nombre) en el panel izquierdo para habilitar el envío.")

            if smtp_cfg and remitentes_df is not None:
                # Build remitente options
                ec = next((c for c in remitentes_df.columns if "mail"  in c.lower()), None)
                cc = next((c for c in remitentes_df.columns if "canal" in c.lower()), None)
                nc = next((c for c in remitentes_df.columns if "nombre" in c.lower()), None)
                rem_opciones = ["— elige destinatario —"]
                rem_map = {}
                if ec and cc:
                    entries = sorted([
                        (f"{r[cc]} — {r[ec]}", (str(r[ec]).strip(), str(r[nc]).strip() if nc else str(r[cc])))
                        for _, r in remitentes_df.iterrows()
                    ], key=lambda x: x[0].lower())
                    for lbl, val in entries:
                        rem_opciones.append(lbl)
                        rem_map[lbl] = val

                # Build pedido labels for multiselect
                pedido_labels = [
                    f"{row['Expedición']} · {row['Canal']} · SKU {row['SKU']} · {row['País']}"
                    for _, row in df_c.iterrows()
                ]

                # Show previous send result
                if '_cancel_msg' in st.session_state:
                    msg_type, msg_text = st.session_state.pop('_cancel_msg')
                    if msg_type == "ok":    st.success(msg_text)
                    elif msg_type == "warn": st.warning(msg_text)
                    else:                    st.error(msg_text)

                # Track sent
                if '_cancel_sent' not in st.session_state:
                    st.session_state['_cancel_sent'] = set()
                sent_set = st.session_state['_cancel_sent']

                # HTML table preview
                rows_html = ""
                for i, (_, row) in enumerate(df_c.iterrows()):
                    sent    = i in sent_set
                    bg      = "#f0fdf4" if sent else "#fff1f2"
                    badge   = "✅ Enviado" if sent else "CANCELAR"
                    badge_bg= "#15803d" if sent else "#b91c1c"
                    rows_html += (
                        f'<tr style="background:{bg}">'
                        f'<td style="padding:6px 10px;border-bottom:1px solid #e8e8e4;font-weight:600">{row["Expedición"]}</td>'
                        f'<td style="padding:6px 10px;border-bottom:1px solid #e8e8e4">{row["Canal"]}</td>'
                        f'<td style="padding:6px 10px;border-bottom:1px solid #e8e8e4"><code>{row["SKU"]}</code></td>'
                        f'<td style="padding:6px 10px;border-bottom:1px solid #e8e8e4">{row["Producto"][:40]}</td>'
                        f'<td style="padding:6px 10px;border-bottom:1px solid #e8e8e4">{row["País"]}</td>'
                        f'<td style="padding:6px 10px;border-bottom:1px solid #e8e8e4">{row["Motivo"][:40]}</td>'
                        f'<td style="padding:6px 10px;border-bottom:1px solid #e8e8e4">'
                        f'<span style="background:{badge_bg};color:#fff;padding:2px 8px;border-radius:8px;font-size:11px;font-weight:700">{badge}</span></td></tr>'
                    )
                st.markdown(
                    f'<table style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:12px">'
                    f'<thead><tr style="background:#141413;color:#fff">'
                    f'<th style="padding:7px 10px;text-align:left">Expedición</th>'
                    f'<th style="padding:7px 10px;text-align:left">Canal</th>'
                    f'<th style="padding:7px 10px;text-align:left">SKU</th>'
                    f'<th style="padding:7px 10px;text-align:left">Producto</th>'
                    f'<th style="padding:7px 10px;text-align:left">País</th>'
                    f'<th style="padding:7px 10px;text-align:left">Motivo</th>'
                    f'<th style="padding:7px 10px;text-align:left">Estado</th>'
                    f'</tr></thead><tbody>{rows_html}</tbody></table>',
                    unsafe_allow_html=True
                )

                with st.form(key="cancel_email_form", clear_on_submit=False):
                    rem_sel     = st.selectbox("Destinatario", rem_opciones)
                    asunto      = st.text_input("Asunto", value=f"Pedidos a cancelar — {len(df_c)} sin sustituto")
                    pedidos_sel = st.multiselect(
                        "Pedidos a incluir en el email",
                        options=pedido_labels,
                        default=pedido_labels,
                    )
                    submitted = st.form_submit_button("📤 Enviar email", type="primary")

                if submitted:
                    email_dest, _ = rem_map.get(rem_sel, (None, None))
                    if not email_dest:
                        st.session_state['_cancel_msg'] = ("warn", "⚠️ Elige un destinatario válido")
                    elif not pedidos_sel:
                        st.session_state['_cancel_msg'] = ("warn", "⚠️ Selecciona al menos un pedido")
                    else:
                        idx_sel  = [i for i, lbl in enumerate(pedido_labels) if lbl in pedidos_sel]
                        df_envio = df_c.iloc[idx_sel].copy()
                        try:
                            send_cancelados_email(email_dest, asunto, df_envio, smtp_cfg)
                            for i in idx_sel:
                                st.session_state['_cancel_sent'].add(i)
                            st.session_state['_cancel_msg'] = ("ok", f"✅ Enviado a {email_dest} ({len(df_envio)} cancelaciones)")
                        except Exception as e:
                            st.session_state['_cancel_msg'] = ("err", f"❌ Error al enviar: {e}")
                    st.rerun()

                if sent_set:
                    if st.button("↩ Limpiar enviados", key="cancel_clear_sent"):
                        st.session_state['_cancel_sent'] = set()
                        st.rerun()
